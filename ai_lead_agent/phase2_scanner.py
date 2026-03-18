"""
╔══════════════════════════════════════════════════════════════════╗
║   RWAtify Phase 2 Scanner                                       ║
║   3 additional lead sources beyond news articles:               ║
║   1. Conference speaker pages (MIPIM, Cityscape, Token2049)     ║
║   2. Job posting signals (hiring = budget + mandate)            ║
║   3. Broader capital-signal Google queries                      ║
║                                                                  ║
║   RUN:                                                          ║
║   python phase2_scanner.py                                      ║
║                                                                  ║
║   Adds leads to the SAME data/rwatify_leads.xlsx                ║
╚══════════════════════════════════════════════════════════════════╝
"""

import os
import sys
import re
import json
import time
import io
import random
import requests
import urllib.parse
from datetime import datetime

# Force UTF-8 for Windows Terminal to prevent UnicodeEncodeError
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.append(current_dir)

from config.rwatify_config import KNOWN_LEADS, COMPETITOR_NAMES, OUTPUT_FILE
from agents.contact_finder import li_company_search, li_person_search

SERPER_API_KEY = os.getenv("SERPER_API_KEY", "")
CLAUDE_API_KEY = os.getenv("CLAUDE_API_KEY", "")

UA_POOL = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36",
]


# ═══════════════════════════════════════════════════════════
#  SOURCE 1 — CONFERENCE SPEAKER PAGES
#  Real estate executives who speak at these events
#  have a mandate and budget — highest quality leads
# ═══════════════════════════════════════════════════════════

CONFERENCE_QUERIES = [
    # Token2049 — RWA / real estate panels
    'site:token2049.com "real estate" speaker 2025',
    '"token2049" "real estate" tokenization speaker panel 2025',

    # Cityscape Global — UAE / MENA real estate
    'site:cityscapeglobal.com speaker "real estate" developer 2025',
    '"Cityscape Global" 2025 speaker "real estate developer" OR "fund manager"',

    # MIPIM — world's largest RE conference
    '"MIPIM" 2025 speaker "real estate" "tokenization" OR "digital" OR "blockchain"',
    'site:mipim.com speaker tokenization 2025',

    # IMN Real Estate
    '"IMN" "real estate" conference speaker "digital" OR "tokenization" 2025',

    # ULI
    '"Urban Land Institute" OR "ULI" speaker "digital real estate" OR "tokenization" 2025',

    # Real Assets conferences
    '"real assets" conference speaker "tokenization" OR "digital investment" 2025',

    # Middle East specific
    '"Arab Real Estate Summit" OR "IndexUAE" speaker tokenization 2025',
    '"Future Investment Initiative" real estate tokenization speaker 2025',
    '"Saudi Real Estate" conference speaker digital tokenization 2025',

    # US specific
    '"Blueprint Vegas" OR "Proptech" conference speaker "tokenization" 2025',
    '"NMHC" OR "multifamily" conference speaker "digital" OR "tokenization" 2025',
]


# ═══════════════════════════════════════════════════════════
#  SOURCE 2 — JOB POSTING SIGNALS
#  Hiring = approved budget + active mandate
#  If a RE company is hiring for tokenization roles → they are building
# ═══════════════════════════════════════════════════════════

JOB_QUERIES = [
    # Direct tokenization/digital asset hires
    '"real estate" "head of tokenization" OR "tokenization lead" job 2025',
    '"real estate" "digital assets director" OR "digital assets head" hiring 2025',
    '"real estate developer" "blockchain engineer" OR "web3" job opening 2025',
    '"real estate fund" "digital transformation" director hiring 2025',
    '"real estate" "investment platform" product manager hiring 2025',
    '"real estate" "head of digital" OR "chief digital officer" job 2025',

    # UAE / MENA job signals
    '"real estate" "digital assets" Dubai job vacancy 2025',
    '"property developer" "blockchain" OR "tokenization" Dubai hiring 2025',

    # Platform build signals from job posts
    '"real estate" "build" "investment platform" engineer hiring',
    '"proptech" "investment platform" "product manager" real estate 2025',
    '"real estate" "investor portal" developer engineer job 2025',
]


# ═══════════════════════════════════════════════════════════
#  SOURCE 3 — BROADER CAPITAL SIGNAL QUERIES
#  Companies with fund/SPV/JV structures that haven't yet
#  tokenized but are RWAtify-ready based on capital complexity
# ═══════════════════════════════════════════════════════════

CAPITAL_SIGNAL_QUERIES = [
    # Fund structure signals
    '"real estate" "Fund II" OR "Fund III" launch 2025',
    '"real estate fund" "first close" OR "final close" 2025',
    '"real estate" "private placement" investors 2025 developer',
    '"real estate" "family office" "co-investment" platform 2025',
    '"real estate" "separate accounts" institutional investors 2025',

    # SPV / JV signals
    '"real estate developer" "joint venture" "capital partner" 2025',
    '"real estate" "SPV" "operating partner" raise 2025',
    '"real estate" "venture" "institutional capital" developer 2025',

    # Platform building signals
    '"real estate" "building our own" "investment platform" 2025',
    '"real estate" "internalize" OR "in-house" "investor management" 2025',
    '"real estate" "replace" "fund administrator" OR "fund admin" 2025',
    '"real estate" "digitize" "cap table" OR "investor reporting" 2025',
    '"real estate" "investor lifecycle" OR "investor operations" platform 2025',

    # Investor portal signals
    '"real estate" "investor portal" "launching" 2025',
    '"real estate" "digital investor" experience platform 2025',

    # MEA / Global signals not covered by main agent
    '"real estate fund" Kenya OR Nigeria OR "South Africa" 2025',
    '"real estate" "family office" Singapore tokenization 2025',
    '"property fund" Australia OR "New Zealand" tokenization digital 2025',
    '"real estate developer" India "fund" "digital" investors 2025',
    '"real estate" "private credit" digital platform 2025',
]


def site_search(q):
    """Helper to return string (handles the f-string in list)."""
    return q


# ═══════════════════════════════════════════════════════════
#  GOOGLE SEARCH (same as main agent)
# ═══════════════════════════════════════════════════════════

def google_search(query: str, num: int = 10) -> list:
    if not SERPER_API_KEY:
        return []
    try:
        time.sleep(random.uniform(0.4, 1.0))
        r = requests.post(
            "https://google.serper.dev/search",
            headers={"X-API-KEY": SERPER_API_KEY, "Content-Type": "application/json"},
            json={"q": query, "num": num, "gl": "us", "hl": "en"},
            timeout=15,
        )
        if r.status_code != 200:
            return []
        results = []
        for item in r.json().get("organic", []):
            url    = item.get("link", "")
            title  = item.get("title", "")
            snippet= item.get("snippet", "")
            source = urllib.parse.urlparse(url).netloc.replace("www.", "")
            if len(title) > 10:
                results.append({"title": title, "url": url, "snippet": snippet, "source": source})
        return results
    except Exception:
        return []


# ═══════════════════════════════════════════════════════════
#  FETCH ARTICLE TEXT
# ═══════════════════════════════════════════════════════════

def fetch_text(url: str) -> str:
    if not url or not url.startswith("http"):
        return ""
    try:
        time.sleep(random.uniform(0.8, 1.5))
        r = requests.get(
            url,
            headers={"User-Agent": random.choice(UA_POOL), "Accept": "text/html,*/*;q=0.8"},
            timeout=12, allow_redirects=True,
        )
        if r.status_code in [403, 404, 429, 451, 503]:
            return ""
        soup = BeautifulSoup(r.text, "lxml")
        for tag in soup(["nav", "footer", "script", "style", "aside", "header"]):
            tag.decompose()
        body = soup.find("article") or soup.find("main") or soup
        text = body.get_text(" ", strip=True)
        return re.sub(r'\s+', ' ', text).strip()[:4000]
    except Exception:
        return ""


# ═══════════════════════════════════════════════════════════
#  CLAUDE QUALIFICATION (simplified for Phase 2)
# ═══════════════════════════════════════════════════════════

PHASE2_PROMPT = """You are qualifying leads for RWAtify — a build-to-own real estate capital infrastructure platform.

RWAtify buyers are: real estate developers, fund managers, family offices, asset operators who raise external capital.
RWAtify is NOT for: tokenization vendors, blockchain consultancies, crypto companies, mortgage brokers, realtors.

READ THIS CAREFULLY:
- SOURCE TYPE is: {source_type}
- If source type is CONFERENCE: The main company is the SPEAKER'S company — extract their name and firm.
- If source type is JOB POSTING: The main company is the one HIRING — extract their name.
- If source type is CAPITAL SIGNAL: The main company is the one raising/managing capital.

ARTICLE TITLE: {title}
SOURCE: {source}
TEXT: {text}

Return JSON only:
{{
  "is_valid_lead": true,
  "is_competitor": false,
  "company_name": "exact name",
  "company_type": "Real Estate Developer | Fund / Investment Manager | Family Office | PropTech | Other",
  "country": "",
  "city": "",
  "company_website": "",
  "decision_maker_name": "",
  "decision_maker_title": "",
  "signal_type": "Conference speaker | Job posting — tokenization role | Capital raise signal | Fund structure",
  "why_rwatify": "one specific sentence",
  "fit_tag": "Strong RWAtify Fit | Potential RWAtify Fit",
  "phase_tag": "Phase 1 — Contact Now | Phase 2 — Nurture | Reject",
  "digital_ready": false,
  "reject_reason": ""
}}"""


def claude_qualify(title: str, source: str, text: str, source_type: str) -> dict:
    if not CLAUDE_API_KEY:
        return {}
    prompt = PHASE2_PROMPT.format(
        source_type=source_type,
        title=title, source=source,
        text=text[:3500],
    )
    try:
        time.sleep(random.uniform(0.5, 1.0))
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": CLAUDE_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-3-haiku-20240307",
                "max_tokens": 600,
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=30,
        )
        if r.status_code != 200:
            return {}
        raw = "".join(b.get("text", "") for b in r.json().get("content", []) if b.get("type") == "text")
        raw = re.sub(r'^```json\s*', '', raw.strip(), flags=re.MULTILINE)
        raw = re.sub(r'^```\s*', '', raw, flags=re.MULTILINE)
        raw = re.sub(r'\s*```$', '', raw, flags=re.MULTILINE)
        return json.loads(raw.strip())
    except Exception:
        return {}


# ═══════════════════════════════════════════════════════════
#  BUILD LEAD RECORD
# ═══════════════════════════════════════════════════════════

def build_record(ai: dict, title: str, url: str, source: str, query: str, source_type: str) -> dict:
    co = (ai.get("company_name") or "").strip()
    dm = (ai.get("decision_maker_name") or "").strip()
    is_known = any(k in co.lower() for k in KNOWN_LEADS) if co else False

    return {
        "Phase Tag":         ai.get("phase_tag", "Phase 2 — Nurture"),
        "Fit Tag":           ai.get("fit_tag", "Potential RWAtify Fit"),
        "Known Lead":        "⭐ Yes — Client List" if is_known else "Discovered",
        "Digital Ready":     "Yes" if ai.get("digital_ready") else "No",
        "Company Name":      co,
        "Company Type":      ai.get("company_type", ""),
        "Country":           ai.get("country", ""),
        "City":              ai.get("city", ""),
        "Company Website":   ai.get("company_website", ""),
        "Company LinkedIn":  li_company_search(co),
        "Decision Maker":    dm,
        "Title":             ai.get("decision_maker_title", ""),
        "Person LinkedIn":   li_person_search(dm, co),
        "Signal Type":       ai.get("signal_type", source_type),
        "Why RWAtify Fits":  ai.get("why_rwatify", ""),
        "Exact Quote":       "",
        "Article Summary":   f"[{source_type}] {title[:120]}",
        "Intent Score":      8 if ai.get("phase_tag") == "Phase 1 — Contact Now" else 5,
        "Reject Reason":     ai.get("reject_reason", ""),
        "News Source":       source,
        "Article Title":     title,
        "Article URL":       url,
        "Article Date":      "",
        "Search Query Used": query,
        "Scraped At":        datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


# ═══════════════════════════════════════════════════════════
#  EXCEL — APPEND TO EXISTING FILE
# ═══════════════════════════════════════════════════════════

TAG_COLORS  = {"Phase 1 — Contact Now": "00C48C", "Phase 2 — Nurture": "F59E0B",
               "Competitor": "EF4444", "Reject": "9CA3AF"}
FIT_COLORS  = {"Strong RWAtify Fit": "065F46", "Potential RWAtify Fit": "92400E"}
BORDER_SIDE = Side(style="thin", color="CBD5E1")

def _bdr():
    return Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)

COLS = [
    ("PHASE", 14, "Phase Tag"), ("FIT TAG", 22, "Fit Tag"), ("KNOWN LEAD", 20, "Known Lead"),
    ("DIGITAL READY", 13, "Digital Ready"), ("COMPANY NAME", 30, "Company Name"),
    ("COMPANY TYPE", 26, "Company Type"), ("COUNTRY", 14, "Country"), ("CITY", 14, "City"),
    ("WEBSITE", 32, "Company Website"), ("DECISION MAKER", 26, "Decision Maker"),
    ("THEIR TITLE", 28, "Title"), ("PERSON LINKEDIN", 46, "Person LinkedIn"),
    ("COMPANY LINKEDIN", 46, "Company LinkedIn"), ("SIGNAL TYPE", 32, "Signal Type"),
    ("WHY RWATIFY FITS", 56, "Why RWAtify Fits"), ("ARTICLE SUMMARY", 56, "Article Summary"),
    ("INTENT SCORE", 12, "Intent Score"), ("NEWS SOURCE", 20, "News Source"),
    ("ARTICLE TITLE", 52, "Article Title"), ("ARTICLE URL", 50, "Article URL"),
    ("SEARCH QUERY", 44, "Search Query Used"), ("SCRAPED AT", 16, "Scraped At"),
]

def append_to_excel(p1_new: list, p2_new: list):
    """Append Phase 2 leads to the existing Excel file, deduplicating by company name."""
    if not p1_new and not p2_new:
        print("  No new leads to add.")
        return

    if not os.path.exists(OUTPUT_FILE):
        print(f"  ⚠️  {OUTPUT_FILE} not found — run main.py first!")
        return

    wb = openpyxl.load_workbook(OUTPUT_FILE)
    nc = len(COLS)

    for tab_name, new_leads in [
        ("⚡ Phase 1 — Contact Now", p1_new),
        ("🟡 Phase 2 — Nurture", p2_new),
    ]:
        if not new_leads:
            continue

        if tab_name not in wb.sheetnames:
            print(f"  ⚠️  Tab '{tab_name}' not found — skipping")
            continue

        ws = wb[tab_name]

        # Get existing company names to deduplicate
        existing_names = set()
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row and row[4]:   # column index 4 = Company Name (0-indexed)
                existing_names.add(str(row[4]).lower().strip())

        added = 0
        for lead in new_leads:
            co = (lead.get("Company Name") or "").lower().strip()
            if co and co in existing_names:
                continue    # skip duplicate

            ri = ws.max_row + 1
            for ci, (_, _, key) in enumerate(COLS, 1):
                val  = str(lead.get(key, "") or "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border    = _bdr()
                cell.fill      = PatternFill("solid", fgColor="F0FDF4" if "Phase 1" in tab_name else "FFF7ED")

                if key in ("Article URL", "Company Website", "Person LinkedIn", "Company LinkedIn"):
                    if val.startswith("http"):
                        cell.hyperlink = val
                        cell.font = Font(color="0563C1", underline="single", name="Arial", size=9)
                        continue
                cell.font = Font(name="Arial", size=9)

            # Phase tag colour
            tc = ws.cell(row=ri, column=1)
            tag = lead.get("Phase Tag", "")
            if tag in TAG_COLORS:
                tc.fill = PatternFill("solid", fgColor=TAG_COLORS[tag])
                tc.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)

            # Known lead gold
            if "Yes" in (lead.get("Known Lead") or ""):
                for ci, (_, _, key) in enumerate(COLS, 1):
                    if key == "Known Lead":
                        kl = ws.cell(row=ri, column=ci)
                        kl.fill = PatternFill("solid", fgColor="FFC107")
                        kl.font = Font(bold=True, color="1A1A1A", name="Arial", size=9)
                        break

            ws.row_dimensions[ri].height = 80
            existing_names.add(co)
            added += 1

        print(f"  [DONE] Added {added} new leads to '{tab_name}'")

    wb.save(OUTPUT_FILE)
    print(f"\n  [SAVE] Saved to {OUTPUT_FILE}")


# ═══════════════════════════════════════════════════════════
#  RUN ALL 3 SOURCES
# ═══════════════════════════════════════════════════════════

def run_source(queries: list, source_type: str, label: str) -> tuple:
    """Run a batch of queries for a specific source type. Returns (p1, p2)."""
    p1, p2 = [], []
    seen_urls = set()
    total_q = len(queries)

    print(f"\n  {'─'*55}")
    print(f"  {label} ({total_q} queries)")
    print(f"  {'─'*55}")

    for i, query in enumerate(queries, 1):
        if not query or not isinstance(query, str):
            continue
        print(f"  [{i:2d}/{total_q}] {query[:65]}", end="  ")
        results = google_search(query, num=10)
        relevant = [r for r in results if r["url"] not in seen_urls]
        print(f"→ {len(relevant)} results")

        for r in relevant:
            seen_urls.add(r["url"])
            title   = r["title"]
            url     = r["url"]
            source  = r["source"]
            snippet = r["snippet"]

            # Fetch full text
            text = fetch_text(url) or snippet

            # Skip pure competitor articles
            combined = (title + " " + text).lower()
            if any(c in combined for c in COMPETITOR_NAMES[:10]) and not any(lead in combined for lead in KNOWN_LEADS):
                continue

            # Claude qualification
            ai = claude_qualify(title, source, text, source_type)
            if not ai or not isinstance(ai, dict):
                continue

            phase = ai.get("phase_tag", "Reject")
            co    = (ai.get("company_name") or "").strip()

            if not ai.get("is_valid_lead") or ai.get("is_competitor") or not co or co == "UNKNOWN":
                print(f"       ❌ Rejected — {ai.get('reject_reason', '')[:50]}")
                continue

            record = build_record(ai, title, url, source, query, source_type)

            if phase == "Phase 1 — Contact Now":
                p1.append(record)
                print(f"       [PHASE 1] Lead: {co}")
            else:
                p2.append(record)
                print(f"       [PHASE 2] Potential: {co}")

    return p1, p2


def main():
    print(f"""
{'═'*60}
  RWAtify Phase 2 Scanner
  Conference Speakers + Job Postings + Capital Signals
  {datetime.now().strftime('%Y-%m-%d %H:%M')}
{'═'*60}
""")

    all_p1, all_p2 = [], []

    # Source 1 — Conference speakers
    p1, p2 = run_source(CONFERENCE_QUERIES, "Conference Speaker", "🎤 Conference Speaker Pages")
    all_p1.extend(p1); all_p2.extend(p2)

    # Source 2 — Job postings
    p1, p2 = run_source(JOB_QUERIES, "Job Posting — Tokenization Role", "💼 Job Posting Signals")
    all_p1.extend(p1); all_p2.extend(p2)

    # Source 3 — Capital signals
    p1, p2 = run_source(CAPITAL_SIGNAL_QUERIES, "Capital Raise Signal", "💰 Capital & Fund Structure Signals")
    all_p1.extend(p1); all_p2.extend(p2)

    print(f"""
{'═'*60}
  Phase 2 Results
{'═'*60}
  ⚡ New Phase 1 leads : {len(all_p1)}
  🟡 New Phase 2 leads : {len(all_p2)}
  Total new leads      : {len(all_p1) + len(all_p2)}
{'═'*60}
""")

    append_to_excel(all_p1, all_p2)

    print(f"""
  [DONE] Open {OUTPUT_FILE} to see all leads.
  Phase 2 leads are now added to the existing tabs.
""")


if __name__ == "__main__":
    main()
