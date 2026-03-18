import pandas as pd
import json
import os
import time
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
import urllib.parse

load_dotenv('.env')
SERPER_API_KEY = os.getenv('SERPER_API_KEY')

def search_serper(query):
    url = "https://google.serper.dev/search"
    payload = json.dumps({"q": query, "num": 3})
    headers = {
        'X-API-KEY': SERPER_API_KEY,
        'Content-Type': 'application/json'
    }
    try:
        response = requests.request("POST", url, headers=headers, data=payload)
        return response.json()
    except Exception as e:
        print(f"Serper error for '{query}': {e}")
        return None

def find_website(company):
    query = f"{company} official website"
    results = search_serper(query)
    if results and 'organic' in results and len(results['organic']) > 0:
        for r in results['organic']:
            link = r.get('link', '')
            # Filter out social media, directories, etc., to find the actual site
            if not any(domain in link for domain in ['linkedin.com', 'facebook.com', 'bloomberg.com', 'zoominfo.com', 'yelp.com', 'wikipedia.org']):
                return link
        return results['organic'][0].get('link', '') # Fallback to first
    return ""

def find_linkedin_profile(company, dm_name, title):
    if not dm_name or dm_name.lower() in ['nan', 'none']: 
        query = f"CEO OR Founder OR Managing Partner {company} LinkedIn"
    else:
        # Sometimes dm_name contains full sentences like "said Susan Tjarksen". Let's clean it roughly.
        clean_dm = dm_name.replace('said ', '').replace('co-founder ', '').strip()
        query = f"{clean_dm} {company} LinkedIn"

    results = search_serper(query)
    if results and 'organic' in results and len(results['organic']) > 0:
        for r in results['organic']:
            link = r.get('link', '')
            if 'linkedin.com/in/' in link:
                return link
    return ""


def process_and_update(json_file, excel_file):
    print("Loading missing leads...")
    with open(json_file, 'r', encoding='utf-8') as f:
        missing_leads = json.load(f)

    print(f"Processing {len(missing_leads)} leads for missing data via Serper...")
    
    updates = []
    
    for idx, lead in enumerate(missing_leads):
        print(f"[{idx+1}/{len(missing_leads)}] Company: {lead['Company']}")
        
        new_website = None
        new_linkedin = None
        
        if lead['Needs_Website']:
            new_website = find_website(lead['Company'])
            print(f"  -> Found Website: {new_website}")

        if lead['Needs_LinkedIn']:
            new_linkedin = find_linkedin_profile(lead['Company'], lead['Current_DM'], lead['Title'])
            print(f"  -> Found LinkedIn: {new_linkedin}")
            
        updates.append({
            'Phase': lead['Phase'],
            'Excel_Row_Index': lead['Excel_Row_Index'],
            'New_Website': new_website,
            'New_LinkedIn': new_linkedin
        })
        time.sleep(0.5) # Basic rate limiting

    print("\nStarting Excel update...")
    # Update Excel using openpyxl to keep formatting
    wb = load_workbook(excel_file)
    ws_phase1 = wb['⚡ Phase 1 — Contact Now']
    ws_phase2 = wb['🟡 Phase 2 — Nurture']
    
    # Based on our previous extraction:
    # Website is column I (9)
    # Person LinkedIn is column L (12)
    WEBSITE_COL = 9
    LINKEDIN_COL = 12
    
    updated_count = 0
    
    for u in updates:
        ws = ws_phase1 if u['Phase'] == 'Phase 1' else ws_phase2
        row_idx = u['Excel_Row_Index']
        
        if u['New_Website']:
            ws.cell(row=row_idx, column=WEBSITE_COL, value=u['New_Website'])
            updated_count += 1
            
        if u['New_LinkedIn']:
            ws.cell(row=row_idx, column=LINKEDIN_COL, value=u['New_LinkedIn'])
            updated_count += 1

    # Save to a new file so the user can verify without loss
    output_file = excel_file.replace('.xlsx', '_updated.xlsx')
    wb.save(output_file)
    print(f"\nDone! Updated {updated_count} cells. Saved to {output_file}")


if __name__ == "__main__":
    process_and_update('data/missing_march14.json', 'data/rwatify_leads.xlsx')
