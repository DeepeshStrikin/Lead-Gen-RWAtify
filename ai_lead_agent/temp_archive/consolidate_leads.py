import os
import pandas as pd
import numpy as np

RAW_DIR = "raw_clients_data"
OUTPUT_FILE = "ALL_CONSOLIDATED_RAW_LEADS.xlsx"

all_leads = []

def find_header_row(df):
    for idx, row in df.iterrows():
        row_str = " ".join([str(val).upper() for val in row.values if pd.notna(val)])
        if "COMPANY NAME" in row_str or "TITLE" in row_str or "DECISION MAKER" in row_str or "ACCOUNT" in row_str:
            return idx
    return -1

for file in os.listdir(RAW_DIR):
    if not file.endswith('.xlsx') or file.startswith("~"):
        continue
    filepath = os.path.join(RAW_DIR, file)
    print(f"Reading: {file}...")
    try:
        xl_raw = pd.ExcelFile(filepath)
        for sheet in xl_raw.sheet_names:
            # Skip noise sheets
            if any(x in sheet.lower() for x in ["reject", "competitor", "summary"]):
                continue
                
            df_raw = xl_raw.parse(sheet, header=None)
            header_idx = find_header_row(df_raw)
            if header_idx == -1:
                # Attempt to use first row if nothing obvious matched
                header_idx = 0
                
            df_raw.columns = df_raw.iloc[header_idx]
            df_raw = df_raw.iloc[header_idx + 1:].reset_index(drop=True)
            
            for _, row in df_raw.iterrows():
                # Helper to gracefully extract column values
                def get_val(*col_names):
                    for col in col_names:
                        matches = [c for c in df_raw.columns if pd.notna(c) and str(c).upper().strip() == col.upper()]
                        if matches:
                            val = row[matches[0]]
                            if pd.notna(val) and str(val).strip() != "":
                                return str(val).strip()
                    return ""
                
                # Extract core data based on expected headers from temp_headers.json
                company = get_val("COMPANY NAME", "Account", "Name", "COMPANY")
                if not company or pd.isna(company) or len(company) < 2:
                    continue
                    
                dm_name = get_val("DECISION MAKER NAME", "Name", "DECISION MAKER")
                title = get_val("TITLE / ROLE", "Title", "ROLE")
                website = get_val("COMPANY WEBSITE", "Website Url", "Website")
                geo = get_val("COUNTRY & CITY", "Geography", "Location", "COUNTRY")
                model = get_val("BUSINESS MODEL")
                portfolio = get_val("PORTFOLIO / PROJECT TYPE")
                li_person = get_val("LINKEDIN — PERSON SEARCH", "LinkedIn URL", "LINKEDIN")
                li_company = get_val("LINKEDIN — COMPANY SEARCH")
                quote = get_val("WHAT THEY SAID (QUOTE)", "EXACT QUOTE (PROOF)")
                why = get_val("WHY RWATIFY IS RELEVANT", "Why RWAtify is Relevant")
                fit = get_val("FIT TAG", "Fit Tag")
                source = get_val("SOURCE", "SOURCE / WEBSITE")
                url = get_val("ARTICLE URL")
                
                all_leads.append({
                    "Company Name": company,
                    "Decision Maker": dm_name,
                    "Title": title,
                    "Website": website,
                    "Location": geo,
                    "Business Model": model,
                    "Portfolio": portfolio,
                    "LinkedIn (Person)": li_person,
                    "LinkedIn (Company)": li_company,
                    "Exact Quote": quote,
                    "Why RWAtify Fit": why,
                    "Fit Tag": fit,
                    "Source File": file,
                    "Source Sheet": sheet,
                    "Article Source": source,
                    "Article URL": url
                })
    except Exception as e:
        print(f"  -> Error on {file}: {e}")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

LEAD_COLUMNS = [
    ("PHASE",               14, "Phase Tag"),
    ("FIT TAG",             22, "Fit Tag"),
    ("KNOWN LEAD",          20, "Known Lead"),
    ("DIGITAL READY",       13, "Digital Ready"),
    ("COMPANY NAME",        30, "Company Name"),
    ("COMPANY TYPE",        26, "Company Type"),
    ("COUNTRY",             14, "Country"),
    ("CITY",                14, "City"),
    ("WEBSITE",             32, "Company Website"),
    ("DECISION MAKER",      26, "Decision Maker"),
    ("THEIR TITLE",         28, "Title"),
    ("PERSON LINKEDIN",     46, "Person LinkedIn"),
    ("COMPANY LINKEDIN",    46, "Company LinkedIn"),
    ("SIGNAL TYPE",         32, "Signal Type"),
    ("WHY RWATIFY FITS",    56, "Why RWAtify Fits"),
    ("EXACT QUOTE",         60, "Exact Quote"),
    ("ARTICLE SUMMARY",     56, "Article Summary"),
    ("INTENT SCORE",        12, "Intent Score"),
    ("NEWS SOURCE",         20, "News Source"),
    ("ARTICLE TITLE",       52, "Article Title"),
    ("ARTICLE URL",         50, "Article URL"),
    ("DATE",                13, "Article Date"),
    ("SEARCH QUERY USED",   44, "Search Query Used"),
    ("SCRAPED AT",          16, "Scraped At"),
]

# Colour palette
C = {
    "p1_hdr": "064E3B",  "p1_row": "F0FDF4",
    "p2_hdr": "78350F",  "p2_row": "FFF7ED",
    "co_hdr": "7F1D1D",  "co_row": "FFF1F2",
    "rj_hdr": "1F2937",  "rj_row": "F9FAFB",
    "tag_p1": "00C48C",  "tag_p2": "F59E0B",
    "tag_co": "EF4444",  "tag_rj": "9CA3AF",
    "note":   "FEF3C7",
}

TAG_COLORS = {
    "Phase 1 — Contact Now": C["tag_p1"],
    "Phase 2 — Nurture":     C["tag_p2"],
    "Competitor":            C["tag_co"],
    "Reject":                C["tag_rj"],
}

FIT_COLORS = {
    "Strong RWAtify Fit":    "065F46",
    "Potential RWAtify Fit": "92400E",
}

def _border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _write_header(ws, row: int, cols: list, bg_hex: str):
    for ci, (h, w, _) in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill      = PatternFill("solid", fgColor=bg_hex)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 36

def _write_data_row(ws, ri: int, lead: dict, cols: list, row_bg: str):
    for ci, (_, _, key) in enumerate(cols, 1):
        val  = str(lead.get(key, "") or "")
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border    = _border()
        cell.fill      = PatternFill("solid", fgColor=row_bg)

        # Hyperlinks
        if key in ("Article URL", "Company Website", "Person LinkedIn", "Company LinkedIn"):
            if val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single", name="Arial", size=9)
                continue

        # Exact quote styling
        if key == "Exact Quote" and val:
            cell.font = Font(italic=True, color="065F46", name="Arial", size=9)
            continue

        cell.font = Font(name="Arial", size=9)

    # Phase tag cell colour
    phase_cell = ws.cell(row=ri, column=1)
    tag = lead.get("Phase Tag", "")
    if tag in TAG_COLORS:
        phase_cell.fill = PatternFill("solid", fgColor=TAG_COLORS[tag])
        phase_cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)

    # Fit tag colour
    if len(cols) > 1:
        fit_cell = ws.cell(row=ri, column=2)
        fit_val  = lead.get("Fit Tag", "")
        if fit_val in FIT_COLORS:
            fit_cell.fill = PatternFill("solid", fgColor=FIT_COLORS[fit_val])
            fit_cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)

    ws.row_dimensions[ri].height = 90

def _make_sheet(wb, name: str, leads: list, cols: list,
                hdr_hex: str, row_hex: str,
                title_txt: str, note_txt: str, active: bool = False):
    ws = wb.active if active else wb.create_sheet(name)
    if active:
        ws.title = name
    ws.sheet_view.showGridLines = False
    nc = len(cols)

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    t = ws["A1"]
    t.value     = title_txt
    t.font      = Font(bold=True, size=11, color="FFFFFF", name="Arial")
    t.fill      = PatternFill("solid", fgColor="1E3A5F")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Note row
    ws.merge_cells(f"A2:{get_column_letter(nc)}2")
    n = ws["A2"]
    n.value     = note_txt
    n.font      = Font(bold=True, size=9, color="7C2D12", name="Arial")
    n.fill      = PatternFill("solid", fgColor=C["note"])
    n.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 22

    # Header row
    _write_header(ws, 3, cols, hdr_hex)

    # Data rows
    for i, lead in enumerate(leads, 4):
        _write_data_row(ws, i, lead, cols, row_hex)

    ws.freeze_panes = "A4"
    if leads:
        ws.auto_filter.ref = f"A3:{get_column_letter(nc)}3"

    return ws

print(f"\n--- CONSOLIDATION RESULTS ---")
df_all = pd.DataFrame(all_leads)
print(f"Total raw leads extracted: {len(df_all)}")

if len(df_all) > 0:
    # 1. Flag records that have a decision maker
    df_all['has_dm'] = df_all['Decision Maker'].apply(lambda x: 0 if not x else 1)
    
    # 2. Sort so that we keep the record WITH the decision maker when deduping
    df_all = df_all.sort_values(by=['has_dm', 'Company Name'], ascending=[False, True])
    
    # 3. Deduplicate based on cleaned company name
    df_all['Company_Norm'] = df_all['Company Name'].str.lower().str.replace(r'[^a-z0-9]', '', regex=True)
    df_clean = df_all.drop_duplicates(subset=['Company_Norm'], keep='first').copy()
    
    print(f"Total unique companies after deduplication: {len(df_clean)}")
    
    # Clean up temp columns
    df_clean = df_clean.drop(columns=['has_dm', 'Company_Norm'])
    
    # Split into Phase 1 (Has DM) and Phase 2 (No DM but valid company)
    phase1_df = df_clean[df_clean["Decision Maker"] != ""]
    phase2_df = df_clean[df_clean["Decision Maker"] == ""]

    print(f"  -> Phase 1 (With Decision Maker): {len(phase1_df)}")
    print(f"  -> Phase 2 (Need to find contact): {len(phase2_df)}")
    
    phase1_list = phase1_df.to_dict('records')
    phase2_list = phase2_df.to_dict('records')

    # Remap Phase Tags appropriately
    for r in phase1_list:
        r["Phase Tag"] = "Phase 1 — Contact Now"
    for r in phase2_list:
        r["Phase Tag"] = "Phase 2 — Nurture"

    wb = openpyxl.Workbook()

    _make_sheet(wb, "⚡ Phase 1 — Contact Now", phase1_list, LEAD_COLUMNS,
                C["p1_hdr"], C["p1_row"],
                "⚡ PHASE 1 — Contact Now  |  Exec quoted on tokenization intent — reach out immediately",
                "Every row has exec quote + proof source. Use WHY RWATIFY FITS as your outreach angle.",
                active=True)

    _make_sheet(wb, "🟡 Phase 2 — Nurture", phase2_list, LEAD_COLUMNS,
                C["p2_hdr"], C["p2_row"],
                "🟡 PHASE 2 — Nurture  |  Tokenization signal found — find exec before outreach",
                "Company shows tokenization interest. Find exec on LinkedIn first, confirm quote, then contact.")
        
    wb.save(OUTPUT_FILE)
    print(f"\n✅ All clean leads have been formatted and saved to: {OUTPUT_FILE}")
else:
    print("No valid leads found to extract.")

