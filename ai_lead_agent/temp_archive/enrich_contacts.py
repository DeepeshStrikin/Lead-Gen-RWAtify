"""
╔══════════════════════════════════════════════════════════════════╗
║   RWAtify Lead Enricher v2                                      ║
║   Fixes existing leads in Excel with real LinkedIn URLs         ║
║   - Saves after each tab (progress never lost)                  ║
║   - Skips broken/encoded URLs (leaves empty instead)            ║
║   - Skips rows already enriched                                 ║
║                                                                  ║
║   RUN:                                                          ║
║   python enrich_contacts.py                                     ║
╚══════════════════════════════════════════════════════════════════╝
"""

import os
import sys
from dotenv import load_dotenv
load_dotenv()

current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.append(current_dir)

import openpyxl
from openpyxl.styles import Font

from config.rwatify_config import OUTPUT_FILE
from agents.contact_finder import find_linkedin_contact

TABS_TO_ENRICH = [
    "⚡ Phase 1 — Contact Now",
    "🟡 Phase 2 — Nurture",
]

# Column positions (1-indexed) — must match LEAD_COLUMNS in main.py
# PHASE(1), FIT TAG(2), KNOWN LEAD(3), DIGITAL READY(4), COMPANY NAME(5),
# COMPANY TYPE(6), COUNTRY(7), CITY(8), WEBSITE(9), DECISION MAKER(10),
# TITLE(11), PERSON LINKEDIN(12), COMPANY LINKEDIN(13) ...
COL_COMPANY_NAME   = 5
COL_WEBSITE        = 9
COL_DECISION_MAKER = 10
COL_TITLE          = 11
COL_PERSON_LI      = 12
COL_COMPANY_LI     = 13


def is_valid_linkedin_url(url: str) -> bool:
    """
    Return True only for clean, real linkedin.com/in/ profile URLs.
    Rejects:
    - URLs with %F0%9D  (mathematical bold unicode — stylised fake names)
    - Non-person slugs  (hub, jobs, company, group etc.)
    - Missing /in/ path
    """
    if not url or "linkedin.com/in/" not in url:
        return False
    # Reject encoded Unicode math characters (stylised bold font in name)
    if "%F0%9D" in url:
        return False
    # Reject slug that is clearly not a person
    bad_slugs = ["hub", "jobs", "company", "group", "page", "official", "recruitment"]
    slug = url.split("/in/")[-1].split("?")[0].lower().rstrip("/")
    if any(slug == b or slug.startswith(b + "-") for b in bad_slugs):
        return False
    return True


def enrich_excel():
    if not os.path.exists(OUTPUT_FILE):
        print(f"❌ {OUTPUT_FILE} not found. Run main.py first!")
        return

    wb = openpyxl.load_workbook(OUTPUT_FILE)
    total_enriched = 0

    for tab_name in TABS_TO_ENRICH:
        if tab_name not in wb.sheetnames:
            print(f"  ⚠️  Tab '{tab_name}' not found — skipping")
            continue

        ws = wb[tab_name]
        print(f"\n  {'═'*55}")
        print(f"  Tab: {tab_name}")
        print(f"  {'═'*55}")

        enriched_this_tab = 0

        for row in ws.iter_rows(min_row=4):   # rows 1-3 = title, note, header
            company = str(row[COL_COMPANY_NAME - 1].value or "").strip()
            existing_li = str(row[COL_PERSON_LI - 1].value or "").strip()

            if not company or company in ("Company Name", "nan", ""):
                continue

            # Skip rows that already have a valid profile URL
            if is_valid_linkedin_url(existing_li):
                print(f"  ✓  {company[:42]} — already enriched")
                continue

            print(f"\n  → Enriching: {company}")

            contact    = find_linkedin_contact(company)
            person_li  = contact.get("linkedin", "")
            name       = contact.get("name", "")
            title      = contact.get("title", "")
            company_li = contact.get("company_linkedin", "")
            website    = contact.get("website", "")

            li_cell = row[COL_PERSON_LI - 1]

            if is_valid_linkedin_url(person_li):
                # ✅ Good URL — write it as a clickable hyperlink
                li_cell.value     = person_li
                li_cell.hyperlink = person_li
                li_cell.font      = Font(color="0563C1", underline="single", name="Arial", size=9)
                print(f"     ✅ Person LinkedIn: {person_li[:65]}")
                enriched_this_tab += 1

                # Write Decision Maker name only if we have a valid profile
                dm_cell = row[COL_DECISION_MAKER - 1]
                if name and not str(dm_cell.value or "").strip():
                    dm_cell.value = name
                    print(f"     ✅ Decision Maker: {name}")

                title_cell = row[COL_TITLE - 1]
                if title and not str(title_cell.value or "").strip():
                    title_cell.value = title
                    print(f"     ✅ Title: {title}")

            else:
                # ❌ Invalid / missing URL — leave Person LinkedIn empty
                li_cell.value = ""
                li_cell.hyperlink = None
                if person_li:
                    print(f"     ⚠️  Invalid URL skipped (encoded/fake) — left empty")
                else:
                    print(f"     ⚠️  No profile found — left empty")

            # Always update Company LinkedIn if empty
            co_li_cell = row[COL_COMPANY_LI - 1]
            if company_li and not str(co_li_cell.value or "").strip():
                co_li_cell.value     = company_li
                co_li_cell.hyperlink = company_li
                co_li_cell.font      = Font(color="0563C1", underline="single", name="Arial", size=9)

            # Always update Website if empty
            web_cell = row[COL_WEBSITE - 1]
            if website and not str(web_cell.value or "").strip():
                web_cell.value     = website
                web_cell.hyperlink = website
                web_cell.font      = Font(color="0563C1", underline="single", name="Arial", size=9)

        total_enriched += enriched_this_tab
        print(f"\n  ✅ {enriched_this_tab} profiles added in '{tab_name}'")

        # ── Save after every tab — progress is never lost ──
        try:
            wb.save(OUTPUT_FILE)
            print(f"  💾 Saved.")
        except PermissionError:
            print(f"\n  ❌ SAVE FAILED — close the Excel file and re-run this script!")
            return

    print(f"""
{'═'*60}
  ✅ ENRICHMENT COMPLETE
{'═'*60}
  Valid LinkedIn profiles added : {total_enriched}
  File saved                    : {OUTPUT_FILE}

  Empty Person LinkedIn = no valid profile found.
  For those rows, use LinkedIn Sales Navigator to find the exec.
{'═'*60}
""")


if __name__ == "__main__":
    enrich_excel()
