import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Input and Output
CORRUPTED_FILE = 'data/rwatify_leads.xlsx'
CLEAN_FILE = 'data/rwatify_leads_RECOVERED.xlsx'

if not os.path.exists(CORRUPTED_FILE):
    print(f"Error: {CORRUPTED_FILE} not found.")
    exit()

def _border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def salvage():
    print(f"Salvaging data from {CORRUPTED_FILE}...")
    xl = pd.ExcelFile(CORRUPTED_FILE)
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    for sheet_name in xl.sheet_names:
        # Clean name: No emojis or special symbols
        clean_sheet_name = sheet_name.encode('ascii', 'ignore').decode('ascii').strip()
        if not clean_sheet_name: clean_sheet_name = "Summary"
        
        print(f"  Processing sheet: {clean_sheet_name}...")
        df = pd.read_excel(CORRUPTED_FILE, sheet_name=sheet_name)
        
        # Create new sheet
        ws = wb.create_sheet(clean_sheet_name)
        ws.sheet_view.showGridLines = False
        
        # Write headers
        headers = df.columns.tolist()
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=str(h))
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            cell.fill = PatternFill("solid", fgColor="1E3A5F")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border()
            ws.column_dimensions[get_column_letter(ci)].width = 25
            
        # Write data
        for ri, row in enumerate(df.values, 2):
            for ci, val in enumerate(row, 1):
                clean_val = str(val) if val is not None else ""
                # Strip long links to avoid OpenXML corruption
                if clean_val.startswith("http") and len(clean_val) > 250:
                    clean_val = clean_val[:250] + "..."
                
                cell = ws.cell(row=ri, column=ci, value=clean_val)
                cell.border = _border()
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Simple link handling
                if clean_val.startswith("http") and "..." not in clean_val:
                    cell.hyperlink = clean_val
                    cell.font = Font(color="0563C1", underline="single", name="Arial", size=9)

    wb.save(CLEAN_FILE)
    print(f"\nSUCCESS! Created clean recovery file: {CLEAN_FILE}")
    print("All leads have been preserved and the formatting is fixed.")

if __name__ == "__main__":
    salvage()
