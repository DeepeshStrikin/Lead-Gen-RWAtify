import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from copy import copy

MASTER_FILE = "data/rwatify_leads.xlsx"
CONSOLIDATED_FILE = "ALL_CONSOLIDATED_RAW_LEADS.xlsx"

print(f"Loading Master Workbook: {MASTER_FILE}...")
wb = load_workbook(MASTER_FILE)

ws_p1 = None
ws_p2 = None

for s in wb.sheetnames:
    if "Phase 1" in s:
        ws_p1 = wb[s]
    elif "Phase 2" in s:
        ws_p2 = wb[s]

print("Loading consolidated lead data...")
df_p1 = pd.read_excel(CONSOLIDATED_FILE, sheet_name="Phase 1 Leads")
df_p2 = pd.read_excel(CONSOLIDATED_FILE, sheet_name="Phase 2 Leads")

def normalize_val(v):
    if pd.isna(v) or v is None:
        return ""
    return str(v).strip()

def append_leads(df, ws, phase_tag):
    # Find the actual headers row in the master sheet (usually row 2 or 3)
    header_row = 1
    for r in range(1, 5):
        if ws.cell(row=r, column=4).value and "COMPANY" in str(ws.cell(row=r, column=4).value).upper():
            header_row = r
            break
            
    # Map header names to column indices so we put data in the EXACT right spot
    col_map = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val:
            col_map[str(val).upper().strip()] = c
            
    # Fallback to standard columns if headers are merged or weird
    col_phase = col_map.get("PHASE TAG", 1)
    col_fit = col_map.get("FIT TAG", 2)
    col_digital = col_map.get("DIGITAL READY?", 3)
    col_company = col_map.get("COMPANY NAME", 4)
    col_website = col_map.get("COMPANY WEBSITE", 5)
    col_location = col_map.get("COUNTRY & CITY", 6)
    col_model = col_map.get("BUSINESS MODEL", 7)
    col_portfolio = col_map.get("PORTFOLIO / PROJECT TYPE", 8)
    col_dm = col_map.get("DECISION MAKER NAME", 9)
    col_title = col_map.get("TITLE / ROLE", 10)
    col_li_person = col_map.get("LINKEDIN — PERSON SEARCH", 11)
    col_li_company = col_map.get("LINKEDIN — COMPANY SEARCH", 12)
    col_quote = col_map.get("EXACT QUOTE (PROOF)", 13)
    col_why = col_map.get("WHY RWATIFY FITS", 14)
    col_signal = col_map.get("SIGNAL TYPE / KEYWORDS", 15)
    col_summary = col_map.get("ARTICLE SUMMARY (AI)", 16)
    col_source = col_map.get("SOURCE", 17)
    col_art_title = col_map.get("ARTICLE TITLE", 18)
    col_art_url = col_map.get("ARTICLE URL", 19)

    # First, build a strict set of ALL companies already in the sheet to prevent duplicates
    existing_companies = set()
    for row in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=row, column=col_company).value
        if val:
            clean_name = str(val).lower().replace(" ", "").replace(".", "").replace(",", "").replace("inc", "").replace("llc", "")
            existing_companies.add(clean_name)

    print(f"  Found {len(existing_companies)} companies already existing in '{ws.title}'.")

    start_row = ws.max_row + 1
    added = 0
    
    hyperlink_font = Font(color="0000FF", underline="single")

    for _, row in df.iterrows():
        comp_raw = normalize_val(row.get('Company Name'))
        if not comp_raw or len(comp_raw) < 2: 
            continue
            
        comp_norm = comp_raw.lower().replace(" ", "").replace(".", "").replace(",", "").replace("inc", "").replace("llc", "")
        
        # Cross dedup with existing master sheet
        if comp_norm in existing_companies:
            print(f"  [Skipping] {comp_raw} is already in the Master Sheet.")
            continue
            
        existing_companies.add(comp_norm)
        
        current_row = start_row + added
        
        # Helper to set value, style, and hyperlink
        def set_cell(c_idx, val, is_link=False):
            cell = ws.cell(row=current_row, column=c_idx, value=val)
            # Copy style from the first data row beneath headers
            template_cell = ws.cell(row=header_row + 1, column=c_idx)
            if template_cell.has_style:
                cell.font = copy(template_cell.font)
                cell.border = copy(template_cell.border)
                cell.fill = copy(template_cell.fill)
                cell.number_format = copy(template_cell.number_format)
                cell.alignment = copy(template_cell.alignment)
            
            # format links properly
            if is_link and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = hyperlink_font
                
        # Inject data based on EXACT column indices from the master sheet headers
        set_cell(col_phase, phase_tag)
        set_cell(col_fit, normalize_val(row.get('Fit Tag')))
        set_cell(col_digital, "")
        set_cell(col_company, comp_raw)
        set_cell(col_website, normalize_val(row.get('Website')), is_link=True)
        set_cell(col_location, normalize_val(row.get('Location')))
        set_cell(col_model, normalize_val(row.get('Business Model')))
        set_cell(col_portfolio, normalize_val(row.get('Portfolio')))
        set_cell(col_dm, normalize_val(row.get('Decision Maker')))
        set_cell(col_title, normalize_val(row.get('Title')))
        set_cell(col_li_person, normalize_val(row.get('LinkedIn (Person)')), is_link=True)
        set_cell(col_li_company, normalize_val(row.get('LinkedIn (Company)')), is_link=True)
        set_cell(col_quote, normalize_val(row.get('Exact Quote')))
        set_cell(col_why, normalize_val(row.get('Why RWAtify Fit')))
        set_cell(col_signal, "")
        set_cell(col_summary, "")
        set_cell(col_source, normalize_val(row.get('Article Source')))
        set_cell(col_art_title, "")
        set_cell(col_art_url, normalize_val(row.get('Article URL')), is_link=True)
            
        added += 1
        
    return added

print("\n--- Appending Phase 1 Leads ---")
added_p1 = append_leads(df_p1, ws_p1, "⚡ PHASE 1 — Contact Now")
print(f"✅ Added {added_p1} NEW unique companies to Phase 1")

print("\n--- Appending Phase 2 Leads ---")
added_p2 = append_leads(df_p2, ws_p2, "🟡 PHASE 2 — Nurture")
print(f"✅ Added {added_p2} NEW unique companies to Phase 2")

print("\nSaving Master File...")
wb.save(MASTER_FILE)
print("SUCCESS! Data aligned perfectly to master headers.")
