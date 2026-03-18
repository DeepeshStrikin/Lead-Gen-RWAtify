"""
╔══════════════════════════════════════════════════════════════════╗
║   RWAtify Lead Agent v2.0 — Main Pipeline                       ║
║   LangGraph orchestration + Colour-coded Excel output           ║
║                                                                  ║
║   HOW THIS WORKS:                                               ║
║   1. fetch_signals   → Serper.dev Google searches (80+ queries) ║
║   2. classify        → Claude AI qualifies each article         ║
║   3. find_contact    → LinkedIn decision maker search           ║
║   4. save_lead       → Append to leads list                     ║
║   5. export_excel    → 5-tab colour-coded Excel output          ║
║                                                                  ║
║   RUN:                                                          ║
║   python main.py                                                ║
║                                                                  ║
║   OUTPUT: data/rwatify_leads.xlsx                               ║
╚══════════════════════════════════════════════════════════════════╝
"""

import pandas as pd
import os
import sys
import io
from datetime import datetime

# Force UTF-8 for Windows Terminal to prevent UnicodeEncodeError
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from typing import TypedDict, List, Dict, Optional, Literal
from langgraph.graph import StateGraph, END

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv
load_dotenv()

# Fix import path
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.append(current_dir)

from source.news_finder import monitor_signals
from agents.intent_classifier import classify_signal, IntentSchema
from agents.contact_finder import find_linkedin_contact, li_person_search, li_company_search
from config.rwatify_config import OUTPUT_FILE, PRE_SEEDED_LEADS, KNOWN_LEADS


# ═══════════════════════════════════════════════════════════
#  GRAPH STATE
# ═══════════════════════════════════════════════════════════

class LeadState(TypedDict):
    signals:        List[Dict]
    current_signal: Optional[Dict]
    intent_result:  Optional[object]    # IntentSchema object
    contact:        Optional[Dict]

    # Lead buckets by phase
    phase1_leads:   List[Dict]
    phase2_leads:   List[Dict]
    competitor_leads: List[Dict]
    rejected_leads: List[Dict]

    # Stats
    query_stats: Dict[str, List[int]]   # query → [total, leads]
    pulse_counter: int


# ═══════════════════════════════════════════════════════════
#  NODE 1 — FETCH SIGNALS
# ═══════════════════════════════════════════════════════════

def seed_known_leads() -> tuple:
    """
    Convert PRE_SEEDED_LEADS from config into full lead records.
    These are companies already identified by the client — always included.
    Returns: (phase1_list, phase2_list)
    """
    p1, p2 = [], []
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    for company in PRE_SEEDED_LEADS:
        co   = company.get("company_name", "")
        tag  = company.get("phase_tag", "Phase 2 — Nurture")

        record = {
            # Phase & qualification
            "Phase Tag":         tag,
            "Fit Tag":           company.get("fit_tag", "Potential RWAtify Fit"),
            "Known Lead":        "⭐ Yes — Client List",
            "Digital Ready":     company.get("digital_ready", "No"),

            # Company
            "Company Name":      co,
            "Company Type":      company.get("company_type", ""),
            "Country":           company.get("country", ""),
            "City":              company.get("city", ""),
            "Company Website":   company.get("company_website", ""),
            "Company LinkedIn":  li_company_search(co),

            # Decision Maker — empty until LinkedIn search is done
            "Decision Maker":    "",
            "Title":             "",
            "Person LinkedIn":   li_person_search("", co),

            # RWAtify context
            "Signal Type":       company.get("signal_type", ""),
            "Why RWAtify Fits":  company.get("why_rwatify", ""),
            "Exact Quote":       "",
            "Article Summary":   f"{co} identified as a high-fit RWAtify prospect based on client research and market intelligence.",
            "Intent Score":      9 if tag == "Phase 1 — Contact Now" else 6,
            "Reject Reason":     "",

            # Source
            "News Source":       "Pre-Seeded — Client Identified",
            "Article Title":     f"{co} — Pre-qualified RWAtify prospect",
            "Article URL":       company.get("company_website", ""),
            "Article Date":      "",
            "Search Query Used": "Pre-seeded (no search required)",
            "Scraped At":        now,
        }

        if tag == "Phase 1 — Contact Now":
            p1.append(record)
        else:
            p2.append(record)

    print(f"  [INFO] Pre-seeded {len(p1)} Phase 1 + {len(p2)} Phase 2 known leads from client list")
    return p1, p2


def fetch_signals(state: LeadState) -> dict:
    """Run all Serper Google searches and inject pre-seeded known leads."""

    # Inject pre-seeded leads first (always guaranteed in output)
    pre_p1, pre_p2 = seed_known_leads()

    # Run internet search
    signals = monitor_signals()

    # Build initial query stats (total articles per query)
    query_stats = {}
    for sig in signals:
        q = sig.get("query", "")
        if q not in query_stats:
            query_stats[q] = [0, 0]
        query_stats[q][0] += 1

    return {
        "signals":          signals,
        "phase1_leads":     pre_p1,   # start with pre-seeded leads
        "phase2_leads":     pre_p2,
        "competitor_leads": [],
        "rejected_leads":   [],
        "query_stats":      query_stats,
    }


# =============================================================
#  NODE 2 - GET NEXT SIGNAL
# =============================================================

def get_next_signal(state: LeadState) -> dict:
    """Pop the next signal from the queue."""
    if not state["signals"]:
        return {}

    signals = state["signals"]
    signal  = signals[0]
    return {
        "current_signal": signal,
        "signals":        signals[1:],
    }


def next_signal_router(state: LeadState) -> str:
    """Route to 'classify' if signals remain, else 'export_excel'."""
    if not state["signals"] and state.get("current_signal") is None:
        return "export_excel"
    return "classify"


# ═══════════════════════════════════════════════════════════
#  NODE 3 — CLASSIFY WITH CLAUDE
# ═══════════════════════════════════════════════════════════

def classify(state: LeadState) -> dict:
    """Qualify the current signal using Claude AI."""
    signal = state.get("current_signal")
    if not signal:
        return state

    print(f"  [AI] Claude qualifying: [{signal.get('domain', 'source')}] {signal.get('title', '')[:50]}...")
    
    try:
        classifier_result = classify_signal(text="", article=signal)
    except Exception as e:
        print(f"  ❌ Classification CRASH: {e}. Recovering with dummy rejection.")
        classifier_result = IntentSchema(
            is_valid_lead=False,
            is_competitor=False,
            reject_reason=f"System error: {str(e)}",
            company_name="ERROR",
            company_type="Error",
            phase_tag="Reject",
            fit_tag="",
            intent="NO",
            urgency=0,
            signal_summary="System encountered an error processing this article.",
        )

    # Increment pulse counter for incremental saving
    pulse = state.get("pulse_counter", 0) + 1

    # Periodically save progress to Excel (Pulse Save)
    if pulse > 0 and pulse % 15 == 0:
        print(f"\n  [SAVE] Pulse Save - Writing {pulse} articles of progress to {OUTPUT_FILE}...")
        export_excel(state)
        print("  [SAVE] Pulse Save Complete.\n")

    return {
        "intent_result": classifier_result,
        "pulse_counter": pulse
    }


def intent_router(state: LeadState) -> str:
    """Route based on Claude's phase tag."""
    result = state.get("intent_result")
    if not result:
        return "skip"

    phase = getattr(result, "phase_tag", "Reject")
    if phase == "Phase 1 — Contact Now":
        return "find_contact"
    elif phase == "Phase 2 — Nurture":
        return "save_lead_p2"
    else:
        return "skip"


# ═══════════════════════════════════════════════════════════
#  NODE 4 — FIND CONTACT (Phase 1 only)
# ═══════════════════════════════════════════════════════════

CONTACT_CACHE: Dict[str, dict] = {}

def find_contact(state: LeadState) -> dict:
    """Find decision maker on LinkedIn for Phase 1 leads."""
    result  = state["intent_result"]
    company = getattr(result, "company_name", "")

    if not company or company == "UNKNOWN":
        return {"contact": {}}

    if company in CONTACT_CACHE:
        return {"contact": CONTACT_CACHE[company]}

    contact = find_linkedin_contact(company)
    CONTACT_CACHE[company] = contact
    return {"contact": contact}


# ═══════════════════════════════════════════════════════════
#  NODE 5a — SAVE PHASE 1 LEAD
# ═══════════════════════════════════════════════════════════

def _build_lead_record(state: LeadState, phase: str) -> dict:
    """Build a full lead record from current state."""
    contact = state.get("contact") or {}
    signal  = state.get("current_signal") or {}
    result  = state.get("intent_result")

    co = getattr(result, "company_name", "") or ""
    dm = contact.get("name", "") or getattr(result, "decision_maker_name", "") or ""
    dm_title = contact.get("title", "") or getattr(result, "decision_maker_title", "") or ""

    # Check if this is a known lead from the client list
    is_known = any(k in co.lower() for k in KNOWN_LEADS) if co else False

    return {
        # Phase & qualification
        "Phase Tag":          phase,
        "Fit Tag":            getattr(result, "fit_tag", ""),
        "Digital Ready":      "Yes" if getattr(result, "digital_ready", False) else "No",
        "Known Lead":         "⭐ Yes — Client List" if is_known else "Discovered",

        # Company
        "Company Name":       co,
        "Company Type":       getattr(result, "company_type", ""),
        "Country":            getattr(result, "country", ""),
        "City":               getattr(result, "city", ""),
        "Company Website":    contact.get("website", "") or getattr(result, "company_website", ""),
        "Company LinkedIn":   contact.get("company_linkedin", "") or li_company_search(co),

        # Decision Maker
        "Decision Maker":     dm,
        "Title":              dm_title,
        "Person LinkedIn":    contact.get("linkedin", "") or contact.get("linkedin_search", "") or li_person_search(dm, co),

        # RWAtify context
        "Signal Type":        getattr(result, "signal_type", ""),
        "Why RWAtify Fits":   getattr(result, "why_rwatify", ""),
        "Exact Quote":        getattr(result, "exact_quote", ""),
        "Article Summary":    getattr(result, "article_summary", getattr(result, "signal_summary", "")),
        "Intent Score":       getattr(result, "urgency", 0),
        "Reject Reason":      getattr(result, "reject_reason", ""),

        # Source
        "News Source":        signal.get("source", ""),
        "Article Title":      signal.get("title", ""),
        "Article URL":        signal.get("url", ""),
        "Article Date":       signal.get("date", ""),
        "Search Query Used":  signal.get("query", ""),
        "Scraped At":         datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


def save_lead_p1(state: LeadState) -> dict:
    """Save a Phase 1 lead (exec quoted — contact now)."""
    result = state.get("intent_result")
    co  = getattr(result, "company_name", "?")
    dm  = state.get("contact", {}).get("name", "—")
    print(f"  [PHASE 1] {co} | {dm}")

    lead = _build_lead_record(state, "Phase 1 — Contact Now")
    leads = list(state.get("phase1_leads", []))
    leads.append(lead)

    # Update query stats
    qs = dict(state.get("query_stats", {}))
    q  = state.get("current_signal", {}).get("query", "")
    if q in qs:
        qs[q][1] += 1

    return {"phase1_leads": leads, "query_stats": qs, "current_signal": None, "contact": {}}


# ═══════════════════════════════════════════════════════════
#  NODE 5b — SAVE PHASE 2 LEAD
# ═══════════════════════════════════════════════════════════

def save_lead_p2(state: LeadState) -> dict:
    """Save a Phase 2 lead (company signal — nurture)."""
    result = state.get("intent_result")
    co = getattr(result, "company_name", "?")
    print(f"  [PHASE 2] {co}")

    lead = _build_lead_record(state, "Phase 2 — Nurture")
    leads = list(state.get("phase2_leads", []))
    leads.append(lead)

    qs = dict(state.get("query_stats", {}))
    q  = state.get("current_signal", {}).get("query", "")
    if q in qs:
        qs[q][1] += 1

    return {"phase2_leads": leads, "query_stats": qs, "current_signal": None, "contact": {}}


# ═══════════════════════════════════════════════════════════
#  NODE 6 — SKIP (Competitor / Reject)
# ═══════════════════════════════════════════════════════════

def skip(state: LeadState) -> dict:
    """Save to appropriate bucket and skip to next signal."""
    result = state.get("intent_result")
    phase  = getattr(result, "phase_tag", "Reject") if result else "Reject"
    co     = getattr(result, "company_name", "?") if result else "?"

    lead = _build_lead_record(state, phase)

    if phase == "Competitor":
        print(f"  🚫 COMPETITOR  |  {co}")
        leads = list(state.get("competitor_leads", []))
        leads.append(lead)
        return {"competitor_leads": leads, "current_signal": None, "contact": {}}
    else:
        reason = getattr(result, "reject_reason", "") if result else ""
        print(f"  ❌ REJECT  |  {reason[:60]}")
        leads = list(state.get("rejected_leads", []))
        leads.append(lead)
        return {"rejected_leads": leads, "current_signal": None, "contact": {}}


# ═══════════════════════════════════════════════════════════
#  EXCEL BUILDER — 5 colour-coded tabs
# ═══════════════════════════════════════════════════════════

LEAD_COLUMNS = [
    # (Header text, width, dict key)
    ("PHASE",               14, "Phase Tag"),
    ("FIT TAG",             22, "Fit Tag"),
    ("KNOWN LEAD",          20, "Known Lead"),
    ("DIGITAL READY",       13, "Digital Ready"),
    ("COMPANY NAME",        30, "Company Name"),
    ("COMPANY TYPE",        26, "Company Type"),
    ("COUNTRY",             14, "Country"),
    ("CITY",                14, "City"),
    ("WEBSITE",             32, "Company Website"),
    ("DECISION MAKER",      26, "Decision Maker"),
    ("THEIR TITLE",         28, "Title"),
    ("PERSON LINKEDIN",     46, "Person LinkedIn"),
    ("COMPANY LINKEDIN",    46, "Company LinkedIn"),
    ("SIGNAL TYPE",         32, "Signal Type"),
    ("WHY RWATIFY FITS",    56, "Why RWAtify Fits"),
    ("EXACT QUOTE",         60, "Exact Quote"),
    ("ARTICLE SUMMARY",     56, "Article Summary"),
    ("INTENT SCORE",        12, "Intent Score"),
    ("NEWS SOURCE",         20, "News Source"),
    ("ARTICLE TITLE",       52, "Article Title"),
    ("ARTICLE URL",         50, "Article URL"),
    ("DATE",                13, "Article Date"),
    ("SEARCH QUERY USED",   44, "Search Query Used"),
    ("SCRAPED AT",          16, "Scraped At"),
]

REJECT_COLUMNS = [
    ("PHASE",          14, "Phase Tag"),
    ("REJECT REASON",  44, "Reject Reason"),
    ("COMPANY NAME",   28, "Company Name"),
    ("NEWS SOURCE",    20, "News Source"),
    ("ARTICLE TITLE",  52, "Article Title"),
    ("ARTICLE URL",    50, "Article URL"),
    ("SEARCH QUERY",   44, "Search Query Used"),
    ("SCRAPED AT",     16, "Scraped At"),
]

# Colour palette
C = {
    "p1_hdr": "064E3B",  "p1_row": "F0FDF4",
    "p2_hdr": "78350F",  "p2_row": "FFF7ED",
    "co_hdr": "7F1D1D",  "co_row": "FFF1F2",
    "rj_hdr": "1F2937",  "rj_row": "F9FAFB",
    "tag_p1": "00C48C",  "tag_p2": "F59E0B",
    "tag_co": "EF4444",  "tag_rj": "9CA3AF",
    "note":   "FEF3C7",
}

TAG_COLORS = {
    "Phase 1 — Contact Now": C["tag_p1"],
    "Phase 2 — Nurture":     C["tag_p2"],
    "Competitor":            C["tag_co"],
    "Reject":                C["tag_rj"],
}

FIT_COLORS = {
    "Strong RWAtify Fit":    "065F46",
    "Potential RWAtify Fit": "92400E",
}

def _border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _write_header(ws, row: int, cols: list, bg_hex: str):
    for ci, (h, w, _) in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill      = PatternFill("solid", fgColor=bg_hex)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 36


def _make_sheet(wb, name_stem, data, columns, hdr_color, row_color, title, subtitle, active=False):
    """Matches sheet by stem (e.g. 'Phase 1') and appends data to the bottom."""
    sheet_name = name_stem
    # Find existing sheet that starts with the name_stem (to handle emojis like '⚡ Phase 1')
    existing_sheets = [s for s in wb.sheetnames if name_stem.lower() in s.lower()]
    
    if existing_sheets:
        ws = wb[existing_sheets[0]]
    else:
        ws = wb.create_sheet(name_stem)
        # Setup basic headers only for new sheets
        ca = ws.cell(row=1, column=1, value=title)
        cb = ws.cell(row=2, column=1, value=subtitle)
        ca.font = Font(bold=True, size=14, color="FFFFFF")
        ca.fill = PatternFill("solid", fgColor="1E3A5F")
        for ci, col in enumerate(columns, 1):
            cell = ws.cell(row=4, column=ci, value=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=hdr_color)

    # Find bottom - start row is max_row + 1
    start_row = ws.max_row + 1
    if ws.max_row < 4: start_row = 5

    # Find Column Index for 'Company Name' dynamically to avoid hardcoding errors
    company_col_idx = 3 # fallback
    for ci in range(1, ws.max_column + 1):
        if ws.cell(row=4, column=ci).value == "Company Name":
            company_col_idx = ci
            break

    # Get existing company names to prevent duplicates
    existing_names = set()
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=company_col_idx).value
        if val: existing_names.add(str(val).lower().strip())

    row_idx = start_row
    for item in data:
        cname = str(item.get("Company Name", "")).lower().strip()
        if not cname or cname in existing_names:
            continue
            
        for ci, col in enumerate(columns, 1):
            val = item.get(col, "")
            if val is None: val = ""
            clean_val = str(val) if not isinstance(val, (int, float)) else val
            
            # Hyperlink protection
            if isinstance(clean_val, str) and clean_val.startswith("http") and len(clean_val) > 250:
                clean_val = clean_val[:250] + "..."

            cell = ws.cell(row=row_idx, column=ci, value=clean_val)
            cell.fill = PatternFill("solid", fgColor=row_color)
            cell.border = _border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name="Arial", size=9)
            
            if isinstance(clean_val, str) and clean_val.startswith("http") and "..." not in clean_val:
                cell.hyperlink = clean_val
                cell.font = Font(color="0563C1", underline="single", name="Arial", size=9)
        
        row_idx += 1
    
    ws.freeze_panes = "A5"
    return ws

def _make_summary_sheet(wb, p1, p2, comp, rej, query_stats):
    """Creates a summary statistics tab. Always overwrites existing Summary."""
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    
    ws = wb.create_sheet("Summary", 0) # Make it the first tab
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 25
    
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    rows = [
        ("RWAtify Lead Agent - Execution Summary", ""),
        ("Last Data Pulse", now),
        ("", ""),
        ("LEAD GENERATION TOTALS (In File)", ""),
        ("⚡ Phase 1 - Contact Now", len(p1)),
        ("🟡 Phase 2 - Nurture", len(p2)),
        ("🚫 Competitors Flagged", len(comp)),
        ("❌ Rejected Entries", len(rej)),
        ("", ""),
        ("SEARCH QUERY PERFORMANCE (Top 15)", ""),
    ]
    
    # Add top queries
    sorted_qs = sorted(query_stats.items(), key=lambda x: -x[1][1])[:15]
    for q, (total, found) in sorted_qs:
        rows.append((f"Query: {q[:60]}...", f"{found} leads found / {total} total articles"))

    for ri, (a, b) in enumerate(rows, 1):
        ws.cell(row=ri, column=1, value=a).font = Font(bold=True if not b else False)
        ws.cell(row=ri, column=2, value=b)
        if "Phase 1" in str(a): ws.cell(row=ri, column=1).font = Font(bold=True, color="065F46")
        if "Phase 2" in str(a): ws.cell(row=ri, column=1).font = Font(bold=True, color="92400E")

def export_excel_file(p1: list, p2: list, comp: list, rej: list, query_stats: dict):
    """Build the final colour-coded Excel with 5 tabs."""
    os.makedirs(os.path.dirname(OUTPUT_FILE) if os.path.dirname(OUTPUT_FILE) else "data", exist_ok=True)

    from openpyxl import load_workbook
    if os.path.exists(OUTPUT_FILE):
        try:
            wb = load_workbook(OUTPUT_FILE)
            print(f"  📂 Loaded existing backup: {OUTPUT_FILE}")
        except Exception as e:
            print(f"  ⚠️ Could not load {OUTPUT_FILE}: {e}. Creating new.")
            wb = openpyxl.Workbook()
    else:
        wb = openpyxl.Workbook()

    # Create sheets or append to existing (matching by name_stem)
    _make_sheet(wb, "Phase 1", p1, LEAD_COLUMNS, C["p1_hdr"], C["p1_row"], "PHASE 1 - Contact Now", "Exec quoted on tokenization intent", active=True)
    _make_sheet(wb, "Phase 2", p2, LEAD_COLUMNS, C["p2_hdr"], C["p2_row"], "PHASE 2 - Nurture", "Tokenization signal found")
    _make_sheet(wb, "Competitor", comp, LEAD_COLUMNS, C["co_hdr"], C["co_row"], "COMPETITORS", "Tokenization infrastructure vendors")
    _make_sheet(wb, "Rejected", rej, REJECT_COLUMNS, C["rj_hdr"], C["rj_row"], "Rejected", "No qualifying RWAtify signal found")

    if "Summary" not in [s for s in wb.sheetnames]:
        _make_summary_sheet(wb, p1, p2, comp, rej, query_stats)

    import time
    for attempt in range(3):
        try:
            wb.save(OUTPUT_FILE)
            break
        except PermissionError:
            print(f"  ⚠️ File locked, retrying in 5s... ({attempt+1}/3)")
            time.sleep(5)
    return len(p1), len(p2), len(comp), len(rej)

def export_excel(state: LeadState) -> dict:
    p1 = state.get("phase1_leads", [])
    p2 = state.get("phase2_leads", [])
    comp = state.get("competitor_leads", [])
    rej = state.get("rejected_leads", [])
    qs = state.get("query_stats", {})

    n1, n2, nc, nr = export_excel_file(p1, p2, comp, rej, qs)

    print(f"\n{'='*65}\n  ✅ DONE — {OUTPUT_FILE}\n{'='*65}")
    print(f"  ⚡ Phase 1 leads added: {n1}\n  🟡 Phase 2 leads added: {n2}\n  🚫 Competitors: {nc}\n  ❌ Rejected: {nr}")
    return state


# ═══════════════════════════════════════════════════════════
#  BUILD LANGGRAPH
# ═══════════════════════════════════════════════════════════

def build_graph():
    builder = StateGraph(LeadState)

    builder.add_node("fetch_signals",   fetch_signals)
    builder.add_node("get_next_signal", get_next_signal)
    builder.add_node("classify",        classify)
    builder.add_node("find_contact",    find_contact)
    builder.add_node("save_lead_p1",    save_lead_p1)
    builder.add_node("save_lead_p2",    save_lead_p2)
    builder.add_node("skip",            skip)
    builder.add_node("export_excel",    export_excel)

    builder.set_entry_point("fetch_signals")

    builder.add_edge("fetch_signals", "get_next_signal")

    builder.add_conditional_edges(
        "get_next_signal",
        next_signal_router,
        {"classify": "classify", "export_excel": "export_excel"},
    )

    builder.add_conditional_edges(
        "classify",
        intent_router,
        {
            "find_contact":  "find_contact",
            "save_lead_p2":  "save_lead_p2",
            "skip":          "skip",
        },
    )

    builder.add_edge("find_contact",   "save_lead_p1")
    builder.add_edge("save_lead_p1",   "get_next_signal")
    builder.add_edge("save_lead_p2",   "get_next_signal")
    builder.add_edge("skip",           "get_next_signal")
    builder.add_edge("export_excel",   END)

    return builder.compile()


graph = build_graph()


# =============================================================
#  ENTRY POINT
# =============================================================

if __name__ == "__main__":
    print(f"""
{'='*65}
  RWAtify Lead Agent v2.0
  Claude + Serper.dev - Entire Internet Google Search
  {datetime.now().strftime('%Y-%m-%d %H:%M')}
{'='*65}
""")
    graph.invoke({
        "signals":          [],
        "current_signal":   None,
        "intent_result":    None,
        "contact":          None,
        "phase1_leads":     [],
        "phase2_leads":     [],
        "competitor_leads": [],
        "rejected_leads":   [],
        "query_stats":      {},
        "pulse_counter":    0,
    })